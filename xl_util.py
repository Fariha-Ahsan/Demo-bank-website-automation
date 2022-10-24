import openpyxl

def rowCount(file,sheet):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    max_rows=sheetname.max_row
    return (max_rows)

def columnCount(file):
    workbook=openpyxl.load_workbook(file)
    sheetname=workbook.sheetnames
    s_n=sheetname[0]
    max_columns=s_n.max_column
    return (max_columns)
def readFile(file,sheet,rows,columns):
    workbook = openpyxl.load_workbook(file)
    sheetname=workbook[sheet]
    return sheetname.cell(column=columns,row=rows).value
